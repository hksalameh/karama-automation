import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import json
import os
import sys
from pathlib import Path
from datetime import datetime
import subprocess
import ctypes
from ctypes import wintypes
import threading
import time
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # Not in a PyInstaller bundle
        base_path = os.path.abspath(os.path.dirname(__file__))
        if os.path.basename(base_path).lower() == '__pycache__':
            base_path = os.path.abspath(os.path.join(base_path, os.pardir))
    return os.path.join(base_path, relative_path)


def writable_app_dir():
    """Return a per-user directory that stays writable after installation."""
    base_dir = os.environ.get('LOCALAPPDATA') or os.environ.get('APPDATA') or os.path.expanduser('~')
    path = os.path.join(base_dir, 'KafalaCompareApp')
    os.makedirs(path, exist_ok=True)
    return path


class AutoUpdateGUI:
    def __init__(self, parent_frame, excel_path, options):
        self.root = parent_frame  # هو Frame وليس Toplevel
        self.excel_path = excel_path
        self.options = options
        self.current_idx = 0
        self.records = []
        self.results = []
        self.paused = True
        self.stop_requested = False
        self.browser_process = None
        self.browser_started = False
        self.processing_started = False  # هل بدأت المعالجة الفعلية (بعد تسجيل الدخول)
        self.browser_ready = False
        self.completed = False  # هل اكتملت العملية
        self.records_loaded = False
        self.build_dir = None
        self.runtime_dir = writable_app_dir()
        self._pre_browser_windows = set()
        
        # شروط التعديل (Checkbox variables)
        self.confirm_var = tk.BooleanVar(value=False)   # تأكيد يدوي
        self.save_var = tk.BooleanVar(value=False)     # تفعيل الحفظ المؤقت
        self.allow_zero_var = tk.BooleanVar(value=True) # السماح بالتصفير
        
        self.setup_ui()
        self.start_update_process()
    
    def setup_ui(self):
        """إعداد واجهة المستخدم"""
        self.root.configure(bg='#f3f4f6')
        
        # شريط العنوان
        header = tk.Frame(self.root, bg='#2f4358', height=80)
        header.pack(fill=tk.X, padx=0, pady=0)
        header.pack_propagate(False)
        
        title = tk.Label(header, text='نافذة مراقبة التعديل التلقائي', 
                        font=('Tahoma', 16, 'bold'), fg='white', bg='#2f4358')
        title.pack(pady=10)
        
        # معلومات الملف
        file_info = tk.Label(header, text=f'الملف: {Path(self.excel_path).name}', 
                            font=('Tahoma', 10), fg='#ddd', bg='#2f4358')
        file_info.pack()
        
        # شريط التقدم
        progress_frame = tk.Frame(self.root, bg='#f3f4f6')
        progress_frame.pack(fill=tk.X, padx=15, pady=10)
        
        self.progress_label = tk.Label(progress_frame, text='جاري التحضير...', 
                                       font=('Tahoma', 10), fg='#111827')
        self.progress_label.pack(anchor='w')
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate', 
                                            length=400, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=5)
        
        # إطار الشروط
        cond_frame = tk.LabelFrame(self.root, text='خيارات التعديل', font=('Tahoma', 10, 'bold'), fg='#2f4358', bg='white', padx=10, pady=8)
        cond_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.confirm_check = tk.Checkbutton(cond_frame, text='تأكيد يدوي قبل كل تعديل', variable=self.confirm_var, bg='white', anchor='e', font=('Tahoma', 10))
        self.confirm_check.pack(anchor='w', padx=5, pady=2)
        self.save_check = tk.Checkbutton(cond_frame, text='تفعيل الحفظ المؤقت بعد كل 15 تعديل', variable=self.save_var, bg='white', anchor='e', font=('Tahoma', 10))
        self.save_check.pack(anchor='w', padx=5, pady=2)
        self.allow_zero_check = tk.Checkbutton(cond_frame, text='السماح بتعديل المبلغ الى 0', variable=self.allow_zero_var, bg='white', anchor='e', font=('Tahoma', 10))
        self.allow_zero_check.pack(anchor='w', padx=5, pady=2)
        
        # حالة المتصفح
        browser_frame = tk.LabelFrame(self.root, text='حالة الموقع', 
                                      font=('Tahoma', 11, 'bold'), 
                                      fg='#2f4358', bg='white')
        browser_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.browser_status_label = tk.Label(browser_frame, text='المتصفح لم يبدأ بعد', 
                                           font=('Tahoma', 10), fg='#111827', bg='white')
        self.browser_status_label.pack(anchor='w', padx=10, pady=5)
        
        self.open_browser_btn = tk.Button(browser_frame, text='جاري التحضير...', 
                                         command=self.start_browser_automation,
                                         bg='#2563eb', fg='white', 
                                         font=('Tahoma', 10, 'bold'),
                                         padx=18, pady=8, cursor='hand2',
                                         state=tk.DISABLED)
        self.open_browser_btn.pack(anchor='w', padx=10, pady=5)
        
        # المحتوى الرئيسي
        main_frame = tk.Frame(self.root, bg='white')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # البطاقة الرئيسية (الحالة الحالية)
        self.card_frame = tk.LabelFrame(main_frame, text='الحالة الحالية', 
                                        font=('Tahoma', 11, 'bold'), 
                                        fg='#2f4358', bg='white')
        self.card_frame.pack(fill=tk.X, pady=10)
        
        # معلومات الحالة
        info_frame = tk.Frame(self.card_frame, bg='#f9fafb')
        info_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # الرقم الوطني والاسم
        top_info = tk.Frame(info_frame, bg='#f9fafb')
        top_info.pack(fill=tk.X, pady=5)
        
        tk.Label(top_info, text='الرقم الوطني:', font=('Tahoma', 10, 'bold'), 
                fg='#2f4358', bg='#f9fafb').pack(side=tk.LEFT, padx=5)
        self.nat_id_label = tk.Label(top_info, text='---', font=('Tahoma', 10), 
                                     fg='#111827', bg='#f9fafb')
        self.nat_id_label.pack(side=tk.LEFT, padx=5)
        
        tk.Label(top_info, text='|', fg='#ddd', bg='#f9fafb').pack(side=tk.LEFT, padx=3)
        
        tk.Label(top_info, text='الاسم:', font=('Tahoma', 10, 'bold'), 
                fg='#2f4358', bg='#f9fafb').pack(side=tk.LEFT, padx=5)
        self.name_label = tk.Label(top_info, text='---', font=('Tahoma', 10), 
                                   fg='#111827', bg='#f9fafb')
        self.name_label.pack(side=tk.LEFT, padx=5)
        
        # المبالغ
        amount_frame = tk.Frame(info_frame, bg='#f9fafb')
        amount_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(amount_frame, text='المبلغ الحالي:', font=('Tahoma', 10, 'bold'), 
                fg='#2f4358', bg='#f9fafb').pack(side=tk.LEFT, padx=5)
        self.old_amount_label = tk.Label(amount_frame, text='---', 
                                        font=('Tahoma', 10, 'bold'), 
                                        fg='#dc2626', bg='#f9fafb')
        self.old_amount_label.pack(side=tk.LEFT, padx=5)
        
        tk.Label(amount_frame, text='→', font=('Tahoma', 12, 'bold'), 
                fg='#9ca3af', bg='#f9fafb').pack(side=tk.LEFT, padx=5)
        
        tk.Label(amount_frame, text='المبلغ الجديد:', font=('Tahoma', 10, 'bold'), 
                fg='#2f4358', bg='#f9fafb').pack(side=tk.LEFT, padx=5)
        self.new_amount_label = tk.Label(amount_frame, text='---', 
                                        font=('Tahoma', 10, 'bold'), 
                                        fg='#16a34a', bg='#f9fafb')
        self.new_amount_label.pack(side=tk.LEFT, padx=5)
        
        # السبب
        reason_frame = tk.Frame(info_frame, bg='#f9fafb')
        reason_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(reason_frame, text='السبب:', font=('Tahoma', 10, 'bold'), 
                fg='#2f4358', bg='#f9fafb').pack(side=tk.LEFT, padx=5)
        self.reason_label = tk.Label(reason_frame, text='---', font=('Tahoma', 9), 
                                     fg='#6b7280', bg='#f9fafb', wraplength=500, justify=tk.LEFT)
        self.reason_label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # الأزرار الرئيسية
        button_frame = tk.Frame(self.card_frame, bg='white')
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        on_back = self.options.get('on_back')
        on_exit = self.options.get('on_exit')

        tk.Button(button_frame, text='↩ العودة للرئيسية', command=on_back if on_back else self.root.destroy, bg='#3b82f6', fg='white', font=('Tahoma', 10, 'bold'), padx=16, pady=8).pack(side=tk.LEFT, padx=5)
        
        self.approve_btn = tk.Button(button_frame, text='✓ موافق', 
                                     command=self.approve_update,
                                     bg='#16a34a', fg='white', 
                                     font=('Tahoma', 10, 'bold'),
                                     padx=20, pady=8, cursor='hand2',
                                     state=tk.DISABLED)
        self.approve_btn.pack(side=tk.LEFT, padx=5)
        
        self.skip_btn = tk.Button(button_frame, text='⊘ تخطي', 
                                 command=self.skip_update,
                                 bg='#f59e0b', fg='white', 
                                 font=('Tahoma', 10, 'bold'),
                                 padx=20, pady=8, cursor='hand2',
                                 state=tk.DISABLED)
        self.skip_btn.pack(side=tk.LEFT, padx=5)
        
        self.cancel_btn = tk.Button(button_frame, text='⛔ إلغاء الكل', 
                                   command=self.cancel_all,
                                   bg='#dc2626', fg='white', 
                                   font=('Tahoma', 10, 'bold'),
                                   padx=20, pady=8, cursor='hand2')
        self.cancel_btn.pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text='✕ خروج', command=on_exit if on_exit else self.root.destroy, bg='#ef4444', fg='white', font=('Tahoma', 10, 'bold'), padx=16, pady=8).pack(side=tk.RIGHT, padx=5)
        
        # سجل التعديلات
        log_frame = tk.LabelFrame(main_frame, text='سجل العمليات', 
                                 font=('Tahoma', 11, 'bold'), 
                                 fg='#2f4358', bg='white')
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, 
                                                 font=('Courier', 9),
                                                 bg='#f3f4f6', fg='#111827',
                                                 wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)
        
        # شريط الحالة السفلي
        status_frame = tk.Frame(self.root, bg='#e5e7eb', height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        status_frame.pack_propagate(False)
        
        self.status_label = tk.Label(status_frame, text='جاري التحضير...', 
                                    font=('Tahoma', 9), fg='#4b5563', bg='#e5e7eb')
        self.status_label.pack(side=tk.LEFT, padx=10, pady=5)

    
    def add_log(self, message, level='INFO'):
        """إضافة رسالة إلى السجل"""
        if threading.current_thread() != threading.main_thread():
            self.root.after(0, lambda: self.add_log(message, level))
            return

        timestamp = datetime.now().strftime('%H:%M:%S')
        colors_map = {
            'INFO': '#111827',
            'SUCCESS': '#16a34a',
            'WARNING': '#f59e0b',
            'ERROR': '#dc2626',
        }
        
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f'[{timestamp}] {message}\n', level)
        self.log_text.tag_config('INFO', foreground=colors_map.get(level, '#111827'))
        self.log_text.tag_config('SUCCESS', foreground=colors_map.get('SUCCESS'))
        self.log_text.tag_config('WARNING', foreground=colors_map.get('WARNING'))
        self.log_text.tag_config('ERROR', foreground=colors_map.get('ERROR'))
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()
    
    def start_update_process(self):
        """بدء عملية التحديث في thread منفصل"""
        thread = threading.Thread(target=self.run_update, daemon=True)
        thread.start()

    def update_browser_status(self, text, level='INFO'):
        if threading.current_thread() != threading.main_thread():
            self.root.after(0, lambda: self.update_browser_status(text, level))
            return
        self.browser_status_label.config(text=text)
        self.add_log(text, level)
    
    def run_update(self):
        """تشغيل سكربت التحديث"""
        try:
            self.root.after(0, lambda: self.open_browser_btn.config(state=tk.DISABLED, text='جاري التحضير...'))
            self.add_log('جاري التحضير للتحديث...', 'INFO')
            self.add_log(f'الملف: {self.excel_path}', 'INFO')
            
            # قراءة البيانات من Excel
            self.load_records_from_excel()
            self.records_loaded = True
            self.add_log(f'✓ تم تجهيز {len(self.records)} سجل للتعديل', 'SUCCESS')
            
            if not self.records:
                self.add_log('لم يتم العثور على أي سجلات قابلة للتعديل في ملف النتيجة', 'WARNING')
                self.root.after(0, self._show_no_records_ready)
                return
            
            time.sleep(1)
            # لا تعرض السجل الأول الآن، انتظر حتى يضغط المستخدم على "بدء المعالجة"
            self.root.after(0, lambda: self.update_initial_state())
        except Exception as e:
            self.add_log(f'خطأ: {str(e)}', 'ERROR')
            self.root.after(0, self._show_no_records_ready)
    
    def update_initial_state(self):
        if not self.records:
            self._show_no_records_ready()
            return
        self.update_browser_status('اضغط زر "فتح الموقع" لبدء المتصفح.', 'INFO')
        self.open_browser_btn.config(state=tk.NORMAL, text='فتح الموقع', command=self.start_browser_automation)
        self.status_label.config(text=f'جاهز للبدء. تم العثور على {len(self.records)} سجل.')

    def _show_no_records_ready(self):
        self.update_browser_status('لا توجد سجلات قابلة للتعديل في ملف النتيجة.', 'WARNING')
        self.open_browser_btn.config(state=tk.NORMAL, text='فتح الموقع', command=self.start_browser_automation)
        self.status_label.config(text='لا توجد سجلات قابلة للتعديل. راجع ورقة "يحتاج تعديل".')

    def load_records_from_excel(self):
        """قراءة البيانات من ملف Excel"""
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            sheet_name = '\u064a\u062d\u062a\u0627\u062c \u062a\u0639\u062f\u064a\u0644'  # يحتاج تعديل
            
            if sheet_name not in wb.sheetnames:
                # جرب الأسماء المختلفة الممكنة
                for sn in wb.sheetnames:
                    if 'تعديل' in sn or 'حتاج' in sn:
                        sheet_name = sn
                        break
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f'لم يتم العثور على ورقة العمل "يحتاج تعديل" - الأوراق المتاحة: {wb.sheetnames}')
            
            ws = wb[sheet_name]
            
            # تجاوز الصفوف الفارغة في البداية
            data_start_row = 1
            for row_idx in range(1, ws.max_row + 1):
                if ws[f'A{row_idx}'].value:
                    data_start_row = row_idx
                    break
            
            # قراءة رؤوس الأعمدة (الصف الأول)
            header_row = data_start_row
            nat_id_col = None
            name_col = None
            old_amount_col = None
            new_amount_col = None
            reason_col = None
            
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(header_row, col_idx).value
                if header:
                    header_str = str(header).strip()
                    if 'وطني' in header_str or 'natId' in header_str.lower():
                        nat_id_col = col_idx
                    elif 'اسم' in header_str and 'الموقع' in header_str:
                        name_col = col_idx
                    elif 'مبلغ' in header_str and 'الموقع' in header_str:
                        old_amount_col = col_idx
                    elif 'مبلغ' in header_str and 'فعلي' in header_str:
                        new_amount_col = col_idx
                    elif 'سبب' in header_str:
                        reason_col = col_idx
            
            # إذا لم نجد أسماء العمود، استخدم الفهارس الافتراضية
            if not nat_id_col:
                nat_id_col = 1
            if not name_col:
                name_col = 2
            if not old_amount_col:
                old_amount_col = 4
            if not new_amount_col:
                new_amount_col = 5
            if not reason_col:
                reason_col = 6
            
            # قراءة البيانات
            self.records = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                nat_id_cell = ws.cell(row_idx, nat_id_col)
                name_cell = ws.cell(row_idx, name_col)
                old_amount_cell = ws.cell(row_idx, old_amount_col)
                new_amount_cell = ws.cell(row_idx, new_amount_col)
                reason_cell = ws.cell(row_idx, reason_col)
                
                nat_id = nat_id_cell.value
                name = name_cell.value
                old_amount = old_amount_cell.value
                new_amount = new_amount_cell.value
                reason = reason_cell.value
                
                if nat_id and (old_amount or new_amount):
                    self.records.append({
                        'natId': str(nat_id).strip(),
                        'name': str(name).strip() if name else '---',
                        'oldAmount': str(old_amount).strip() if old_amount else '',
                        'newAmount': str(new_amount).strip() if new_amount else '',
                        'reason': str(reason).strip() if reason else '',
                    })
            
            wb.close()
        except Exception as e:
            self.add_log(f'خطأ في قراءة الملف: {str(e)}', 'ERROR')
            raise
    
    def start_browser_automation(self):
        """بدء سكربت Playwright في thread منفصل"""
        try:
            if not self.records_loaded:
                messagebox.showwarning('تنبيه', 'ما زال التطبيق يقرأ ملف النتيجة. انتظر لحظات ثم أعد المحاولة.')
                return
            if not self.records:
                self.add_log('⚠️  لا توجد سجلات قابلة للتعديل في ملف النتيجة، سيتم فتح الموقع فقط.', 'WARNING')

            if self.browser_started:
                self.add_log('المتصفح قيد التشغيل بالفعل.', 'WARNING')
                return

            # Reset end-of-run state so the same session can be started again.
            self.completed = False
            self.stop_requested = False
            self.processing_started = False
            self.current_idx = 0

            if getattr(sys, 'frozen', False):
                # The application is frozen (e.g., by PyInstaller)
                base_dir = os.path.dirname(sys.executable)
            else:
                # The application is running as a script
                base_dir = os.path.dirname(os.path.abspath(__file__))

            script_path = resource_path('auto_update_from_diff.js')
            
            if not os.path.exists(script_path):
                raise FileNotFoundError(f'سكربت التحديث غير موجود: {script_path}')
            
            # إنشاء ملف مؤقت للبيانات للتواصل بين GUI و Node.js.
            # Program Files is read-only for normal users, so keep runtime
            # files in the user's local app data directory.
            self.runtime_dir = writable_app_dir()
            self.control_file = os.path.join(self.runtime_dir, f'_control_{int(time.time())}.json')
            
            # تمرير مسار Excel عن طريق متغير البيئة (ملف مؤقت يحتوي على المسار)
            env = os.environ.copy()
            self.path_file = os.path.join(self.runtime_dir, f'_karama_path_{int(time.time())}.txt')
            with open(self.path_file, 'w', encoding='utf-8') as f:
                f.write(self.excel_path)
            env['KARAMA_PATH_FILE'] = self.path_file
            env['KARAMA_CONTROL_FILE'] = self.control_file
            env['KARAMA_OUTPUT_DIR'] = self.runtime_dir
            
            # بناء أوامر سطر الأوامر
            # In frozen mode, node.exe is at _MEIPASS root (per fixed.spec layout)
            # In script mode, it's in the KafalaCompareApp_build folder
            if getattr(sys, 'frozen', False):
                node_path_bundled = resource_path('node.exe')
                self.build_dir = resource_path('.')
            else:
                node_path_bundled = os.path.join(base_dir, 'KafalaCompareApp_build', 'node.exe')
                self.build_dir = os.path.join(base_dir, 'KafalaCompareApp_build')

            node_candidates = [
                node_path_bundled, # First, check for the bundled version
                'node', # Fallback to system PATH
            ]
            node_path = None
            for candidate in node_candidates:
                if self._find_executable(candidate):
                    node_path = candidate
                    break
            if not node_path:
                raise FileNotFoundError('تعذر العثور على node.exe أو node في PATH. تأكد من تثبيت Node.js.')
            
            cmd = [node_path, script_path, '--from-gui']
            # قراءة الشروط من واجهة المستخدم
            # The GUI itself controls manual approval before sending each record.
            # Keep the Node side non-interactive so it processes exactly the
            # command the GUI already approved.
            cmd.append('--yes')
            if self.allow_zero_var.get():
                cmd.append('--allow-zero')
            if self.save_var.get():
                cmd.append('--save')
            
            self.add_log(f'cmd: {cmd}', 'INFO')
            self.update_browser_status('جاري تشغيل المتصفح...', 'INFO')
            self.open_browser_btn.config(state=tk.DISABLED)
            
            # التقاط النوافذ الموجودة قبل تشغيل المتصفح
            self._capture_existing_browser_windows()

            # تشغيل السكربت في thread
            self.browser_thread = threading.Thread(
                target=self._run_browser_process,
                args=(cmd, env),
                daemon=True
            )
            self.browser_thread.start()
            self.add_log('✓ جاري بدء عملية المتصفح...', 'SUCCESS')

            # تعطيل الخيارات بعد بدء التشغيل
            self.confirm_check.config(state=tk.DISABLED)
            self.save_check.config(state=tk.DISABLED)
            self.allow_zero_check.config(state=tk.DISABLED)
            self.add_log('🔒 تم تثبيت خيارات التعديل. لا يمكن تغييرها الآن.', 'INFO')
        except Exception as e:
            self.open_browser_btn.config(state=tk.NORMAL)
            self.update_browser_status(f'خطأ في فتح المتصفح: {str(e)}', 'ERROR')
            raise
    
    def _find_executable(self, candidate):
        """التحقق من وجود ملف تنفيذي محلي أو على PATH."""
        from shutil import which

        if not candidate:
            return False

        # If the caller already gave us an absolute path, use it directly.
        if os.path.isabs(candidate):
            return os.path.isfile(candidate)

        # Otherwise fall back to PATH lookup.
        return which(candidate) is not None
    
    def _run_browser_process(self, cmd, env):
        """تشغيل عملية المتصفح"""
        try:
            self.browser_process = subprocess.Popen(
                cmd,
                env=env, cwd=self.runtime_dir, creationflags=subprocess.CREATE_NO_WINDOW,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                stdin=subprocess.DEVNULL,
                text=True,
                encoding='utf-8',
                errors='replace'
            )
            self.root.after(0, self._on_browser_started)

            # قراءة stdout والكشف عن إشارة الإكمال
            if self.browser_process.stdout:
                for line in self.browser_process.stdout:
                    s = line.strip()
                    if '__DONE__' in s:
                        self.add_log('DONE signal detected', 'SUCCESS')
                        self.root.after(0, lambda: self._on_browser_finished(0))
                        break
                    if s:
                        self.add_log(s, 'INFO')
            if self.browser_process.stderr:
                for line in self.browser_process.stderr:
                    s = line.strip()
                    if s:
                        self.add_log(s, 'ERROR')

            return_code = self.browser_process.wait()
            self.root.after(0, lambda: self._on_browser_finished(return_code))
        except Exception as e:
            self.root.after(0, lambda: self._on_browser_failed(str(e)))

    def _on_browser_started(self):
        self.browser_started = True
        # بعد نجاح فتح المتصفح، اجعل الزر ينتقل دائمًا إلى "بدء المعالجة".
        if not self.records:
            self.update_browser_status('✓ تم فتح المتصفح. لا توجد سجلات قابلة للتعديل، لذلك سيبقى البدء معطلاً عند الضغط.', 'WARNING')
            self.open_browser_btn.config(
                text='🚀 بدء المعالجة',
                command=self._send_start_signal,
                state=tk.NORMAL
            )
            self.status_label.config(text='المتصفح مفتوح، لكن لا توجد سجلات قابلة للتعديل.')
            self._position_main_window_right()
            self.root.after(2000, self._try_position_browser_left, 0)
            return

        self.update_browser_status('✓ تم فتح المتصفح. سجل الدخول ثم اضغط "بدء المعالجة".', 'SUCCESS')
        self.open_browser_btn.config(
            text='🚀 بدء المعالجة',
            command=self._send_start_signal,
            state=tk.NORMAL
        )
        self.status_label.config(text='في انتظار إشارة البدء بعد تسجيل الدخول...')

        # تقسيم الشاشة: البرنامج يميناً والمتصفح يساراً
        self._position_main_window_right()
        self.root.after(2000, self._try_position_browser_left, 0)

    def _send_start_signal(self):
        if self.processing_started:
            return
        if not self.records:
            messagebox.showwarning('تنبيه', 'لا توجد سجلات قابلة للتعديل.')
            self._show_no_records_ready()
            return
        self.processing_started = True
        self.send_command_to_browser({'action': 'start'})
        self.add_log('إشارة البدء أُرسلت، في انتظار أول سجل...', 'INFO')
        self.open_browser_btn.config(state=tk.DISABLED)
        # Now show the first record and enable approve/skip
        self.root.after(100, self.show_next_record)

    def _on_browser_finished(self, return_code):
        if self.completed:
            return
        if return_code == 0:
            self.add_log('✅ انتهت عملية المتصفح بنجاح.', 'SUCCESS')
        else:
            self.add_log(f'⚠️ انتهت عملية المتصفح برمز: {return_code}', 'ERROR')
        self.open_browser_btn.config(state=tk.DISABLED)
        self.browser_started = False
        self.processing_started = False
        self.stop_requested = True
        self.finish_update()

    def _on_browser_failed(self, error_message):
        self.update_browser_status(f'فشل المتصفح: {error_message}', 'ERROR')
        self.open_browser_btn.config(state=tk.NORMAL, text='فتح الموقع', command=self.start_browser_automation)
        self.browser_started = False
        self.processing_started = False
        if not self.stop_requested:
            self.finish_update()

    def show_next_record(self):
        """عرض الحالة التالية"""
        if self.stop_requested or self.current_idx >= len(self.records):
            if not self.completed:
                self.finish_update()
            return

        record = self.records[self.current_idx]

        self.nat_id_label.config(text=record['natId'])
        self.name_label.config(text=record['name'])
        self.old_amount_label.config(text=record['oldAmount'] or '(فارغ)')
        self.new_amount_label.config(text=record['newAmount'])
        self.reason_label.config(text=record['reason'])

        progress = ((self.current_idx) / len(self.records)) * 100
        self.progress_bar['value'] = progress

        self.status_label.config(text=f'[{self.current_idx + 1}/{len(self.records)}] {record["natId"]}')

        # إذا لم يكن التأكيد اليدوي مفعلاً، قم بالمعالجة تلقائياً بعد فترة قصيرة
        if not self.confirm_var.get():
            self.add_log(f'➤ [تلقائي] {record["natId"]} - {record["name"]}', 'INFO')
            self.root.after(500, self.approve_update)
        else:
            # إذا كان التأكيد مفعلاً، انتظر تفاعل المستخدم
            self.approve_btn.config(state=tk.NORMAL)
            self.skip_btn.config(state=tk.NORMAL)
            self.add_log(f'➤ [{self.current_idx + 1}/{len(self.records)}] {record["natId"]} - {record["name"]}', 'INFO')
            self.add_log(f'  التعديل المطلوب: {record["oldAmount"] or "فارغ"} → {record["newAmount"]}', 'INFO')

    def approve_update(self):
        if not self.browser_started or not self.processing_started:
            messagebox.showwarning('خطأ', 'يجب بدء المعالجة أولاً.')
            return

        record = self.records[self.current_idx]
        self.add_log(f'أمر معالجة أُرسل لـ: {record["natId"]}', 'SUCCESS')
        
        command = {
            'action': 'process',
            'record': record,
            'index': self.current_idx,
            'total': len(self.records),
        }
        self.send_command_to_browser(command)
        
        self.results.append({'natId': record['natId'], 'status': 'approved_for_processing'})
        self.approve_btn.config(state=tk.DISABLED)
        self.skip_btn.config(state=tk.DISABLED)
        self.status_label.config(text=f'جاري التعديل في الموقع لـ {record["natId"]}...')
        
        self.current_idx += 1
        # Use a small delay to allow the backend to process before showing the next item
        self.root.after(500, self.show_next_record)

    def skip_update(self):
        """تخطي التعديل"""
        if not self.browser_started or not self.processing_started:
            messagebox.showinfo('تنبيه', 'يجب بدء المعالجة أولاً بالضغط على "موافق".')
            return

        record = self.records[self.current_idx]
        self.add_log(f'⊘ تم التخطي: {record["natId"]}', 'WARNING')
        self.results.append({'natId': record['natId'], 'status': 'skipped'})
        
        # Send skip command to backend so it knows to expect the next record
        self.send_command_to_browser({
            'action': 'skip',
            'record': record
        })
        
        self.approve_btn.config(state=tk.DISABLED)
        self.skip_btn.config(state=tk.DISABLED)
        
        self.current_idx += 1
        self.root.after(100, self.show_next_record) # Short delay before showing next

    def send_command_to_browser(self, command):
        """إرسال أمر كامل إلى سكربت المتصفح عن طريق ملف JSON"""
        try:
            # Add a timestamp to the command
            command['timestamp'] = int(time.time())
            
            if not hasattr(self, 'control_file'):
                self.add_log('⚠️  ملف التحكم غير جاهز.', 'ERROR')
                return
            
            with open(self.control_file, 'w', encoding='utf-8') as f:
                json.dump(command, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.add_log(f'خطأ في إرسال الأمر: {str(e)}', 'ERROR')
    
    def cancel_all(self):
        """إلغاء جميع العمليات"""
        if messagebox.askyesno('تأكيد الإلغاء', 'هل تريد إلغاء جميع التعديلات؟'):
            self.stop_requested = True
            if self.browser_process and self.browser_process.poll() is None:
                self.send_command_to_browser({'action': 'shutdown'})
            self.add_log('⛔ تم إلغاء العملية', 'ERROR')
            self.finish_update()
    
    def finish_update(self):
        """إنهاء العملية - البقاء في نفس الشاشة"""
        if self.completed:
            return
        self.completed = True
        
        self.approve_btn.config(state=tk.DISABLED, text='✓ تم الانتهاء')
        self.skip_btn.config(state=tk.DISABLED)
        self.open_browser_btn.config(state=tk.NORMAL, text='فتح الموقع', command=self.start_browser_automation)
        
        # # حفظ النتائج - تم التعطيل لتجنب إنشاء ملف مضلل
        # result_file = f'auto_update_result_{int(time.time())}.json'
        # with open(result_file, 'w', encoding='utf-8') as f:
        #     json.dump(self.results, f, ensure_ascii=False, indent=2)
        
        # تحديث الحالة في نفس الشاشة
        self.status_label.config(
            text=f'✅ تم الانتهاء!'
        )
        self.progress_label.config(text='🔔 اضغط حفظ مؤقت في الموقع لحفظ التعديلات')
        self.progress_bar['value'] = 100
        
        self.add_log(f'✅ انتهت العملية.', 'SUCCESS')
        self.add_log(f'📊 تم إنشاء تقرير مفصل في: {self.runtime_dir}', 'INFO')
        self.add_log('🔔 تذكر: اضغط حفظ مؤقت في الموقع لحفظ آخر التعديلات!', 'WARNING')

        # إعادة تفعيل الخيارات لجولة جديدة
        self.confirm_check.config(state=tk.NORMAL)
        self.save_check.config(state=tk.NORMAL)
        self.allow_zero_check.config(state=tk.NORMAL)


    def _capture_existing_browser_windows(self):
        """"Capture the set of existing Chrome browser windows before launching."""
        try:
            user32 = ctypes.windll.user32
            self._pre_browser_windows = set()
            def enum_cb(hwnd, lparam):
                if user32.IsWindowVisible(hwnd):
                    cls = ctypes.create_unicode_buffer(256)
                    user32.GetClassNameW(hwnd, cls, 256)
                    if cls.value == 'Chrome_WidgetWin_1':
                        self._pre_browser_windows.add(hwnd)
                return True
            WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            user32.EnumWindows(WNDENUMPROC(enum_cb), 0)
        except Exception as e:
            self.add_log(f'Window capture warning: {e}', 'WARNING')

    def _position_main_window_right(self):
        """"Position the main app window on the right half of the screen."""
        try:
            user32 = ctypes.windll.user32
            screen_w = user32.GetSystemMetrics(0)
            screen_h = user32.GetSystemMetrics(1)
            main_root = self.root.winfo_toplevel()
            hwnd = int(main_root.winfo_id())
            half_w = screen_w // 2
            margin = 8
            user32.ShowWindow(hwnd, 1)
            user32.MoveWindow(hwnd, half_w + margin, margin, half_w - 2 * margin, screen_h - 2 * margin, True)
        except Exception as e:
            self.add_log(f'Window position warning: {e}', 'WARNING')

    def _try_position_browser_left(self, attempt):
        """"Try to find and position the browser window on the left half."""
        if attempt >= 10:
            return
        hwnd = self._find_new_browser_window()
        if hwnd:
            try:
                user32 = ctypes.windll.user32
                screen_w = user32.GetSystemMetrics(0)
                screen_h = user32.GetSystemMetrics(1)
                half_w = screen_w // 2
                margin = 8
                user32.ShowWindow(hwnd, 1)
                user32.MoveWindow(hwnd, margin, margin, half_w - 2 * margin, screen_h - 2 * margin, True)
            except Exception as e:
                self.add_log(f'Browser position warning: {e}', 'WARNING')
        else:
            self.root.after(800, self._try_position_browser_left, attempt + 1)

    def _find_new_browser_window(self):
        """"Find a Chrome browser window that wasn't present before launch."""
        try:
            user32 = ctypes.windll.user32
            found = []
            def enum_cb(hwnd, lparam):
                if user32.IsWindowVisible(hwnd):
                    cls = ctypes.create_unicode_buffer(256)
                    user32.GetClassNameW(hwnd, cls, 256)
                    if cls.value == 'Chrome_WidgetWin_1':
                        title = ctypes.create_unicode_buffer(512)
                        user32.GetWindowTextW(hwnd, title, 512)
                        rect = wintypes.RECT()
                        user32.GetWindowRect(hwnd, ctypes.byref(rect))
                        area = (rect.right - rect.left) * (rect.bottom - rect.top)
                        if area > 20000 and hwnd not in self._pre_browser_windows:
                            found.append((hwnd, area, title.value))
                return True
            WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            user32.EnumWindows(WNDENUMPROC(enum_cb), 0)
            if found:
                found.sort(key=lambda x: -x[1])
                return found[0][0]
            return None
        except Exception:
            return None


def main():
    root = tk.Tk()
    
    # مثال على الاستخدام
    excel_path = 'result.xlsx'
    options = {'confirm': True, 'save': False, 'allow_zero': False}
    
    app = AutoUpdateGUI(root, excel_path, options)
    root.mainloop()


if __name__ == '__main__':
    main()
